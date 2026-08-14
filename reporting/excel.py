"""Excel workbook writer for a collection report.

Uses the same section tables as PSV. Adds charts where a numeric series exists.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reporting.sections import REPORT_SECTIONS, is_unavailable

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")


def _headers(rows: list[dict[str, Any]]) -> list[str]:
    headers: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in headers:
                headers.append(key)
    return headers


def _write_sheet(ws, rows: list[dict[str, Any]], *, title: str) -> None:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if not rows:
        ws["A3"] = "Not available for this historical run"
        return
    if is_unavailable(rows):
        ws["A3"] = "status"
        ws["B3"] = "message"
        ws["A3"].font = Font(bold=True)
        ws["B3"].font = Font(bold=True)
        ws["A4"] = rows[0].get("status", "")
        ws["B4"] = rows[0].get("message", "Not available for this historical run")
        return
    headers = _headers(rows)
    header_row = 3
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, idx, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(rows, start=header_row + 1):
        for c_idx, header in enumerate(headers, start=1):
            ws.cell(r_idx, c_idx, row.get(header, "N/A"))
    for idx, header in enumerate(headers, start=1):
        width = min(max(len(str(header)) + 2, 14), 48)
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(rows)}"


def _add_chart(ws, rows: list[dict[str, Any]], value_col: str) -> None:
    if is_unavailable(rows) or not rows:
        return
    headers = _headers(rows)
    if value_col not in headers:
        return
    cat_col = 1
    val_col = headers.index(value_col) + 1
    numeric_rows = 0
    for row in rows:
        try:
            float(row.get(value_col))
            numeric_rows += 1
        except (TypeError, ValueError):
            continue
    if numeric_rows < 1:
        return
    header_row = 3
    last_row = header_row + len(rows)
    chart = BarChart()
    chart.type = "col"
    chart.y_axis.title = value_col
    data = Reference(ws, min_col=val_col, min_row=header_row, max_row=last_row)
    cats = Reference(ws, min_col=cat_col, min_row=header_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 15
    chart.height = 8
    ws.add_chart(chart, "H3")


def write_excel(path: Path, tables: dict[str, list[dict[str, Any]]]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    first = True
    for _heading, key, sheet_title, chart_col in REPORT_SECTIONS:
        if first:
            ws = workbook.active
            ws.title = sheet_title
            first = False
        else:
            ws = workbook.create_sheet(sheet_title)
        rows = list(tables.get(key) or [])
        _write_sheet(ws, rows, title=sheet_title)
        if chart_col:
            _add_chart(ws, rows, chart_col)
    workbook.save(path)
    return path
