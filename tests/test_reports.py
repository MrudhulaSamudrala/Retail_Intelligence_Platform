"""Report generation and discovery tests — no live collection."""

from __future__ import annotations

import asyncio
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import AsyncMock, patch

import pytest
from sqlalchemy import create_engine, event, select
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from collector.orchestration.config import STATUS_SUCCESS, OrchestrationConfig
from collector.orchestration.runner import ProductionRunner
from collector.orchestration.steps import StepResult
from dashboard.queries.reports import (
    attach_run_metadata,
    discover_reports,
    latest_report,
    metrics_from_steps,
    split_latest_and_previous,
)
from dashboard.views.collection_status import _component_status, _normalize
from database.models import (
    Base,
    CollectionRun,
    CollectionRunStep,
    PriceHistory,
    Product,
    ProductSnapshot,
    RetailerAudit,
)
from reporting.excel import write_excel
from reporting.generate import generate_historical_production_reports, generate_reports_for_run
from reporting.paths import paths_for_run
from reporting.psv import write_psv
from reporting.sections import REPORT_SECTIONS
from reporting.run_scope import (
    list_production_run_ids,
    observation_run_ids_for_run,
    parse_run_id_list,
)


@pytest.fixture()
def session() -> Session:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        future=True,
    )
    @event.listens_for(engine, "connect")
    def _fk_on(dbapi_connection, connection_record):  # type: ignore[no-untyped-def]
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()

    Base.metadata.create_all(bind=engine)
    factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    sess = factory()
    try:
        yield sess
    finally:
        sess.close()
        engine.dispose()


def test_psv_and_excel_writers(tmp_path: Path) -> None:
    tables = {
        "executive": [{"metric": "products_observed", "value": 200}],
        "details": [
            {"metric": "collection_run_id", "value": 8},
            {"metric": "status", "value": "PARTIAL"},
        ],
        "coverage": [{"retailer": "newegg", "country": "US", "products": 100, "status": "SUCCESS"}],
        "status": [{"component": "newegg", "status": "SUCCESS", "records": 100}],
        "shelf": [{"brand": "Intel", "product_count": 10, "universe_size": 20, "share_percent": 50}],
        "visibility": [
            {"brand": "Intel", "appearances": 5, "share_of_voice": 40, "coverage_status": "exact"}
        ],
        "pricing": [{"brand": "Intel", "observations": 10, "average_price": 999.0, "currency": "USD"}],
        "promotions": [
            {
                "brand": "Intel",
                "products_with_promotion": 2,
                "discounted_products": 3,
                "average_discount": 10,
            }
        ],
        "promotion_details": [{"product_id": 1, "brand": "Intel", "current_price": 900}],
        "compliance": [
            {
                "brand": "Intel",
                "s1": 94,
                "s2": 94,
                "p1": 100,
                "p2": 94,
                "p3": 93,
                "p4": 94,
                "p5": 89,
                "overall": 99,
            },
            {
                "brand": "AMD",
                "s1": 88,
                "s2": 88,
                "p1": 94,
                "p2": 88,
                "p3": 65,
                "p4": 88,
                "p5": 82,
                "overall": 100,
            },
        ],
        "banners": [
            {"status": "NOT_AVAILABLE", "message": "Not available for this historical run"}
        ],
        "quality": [{"attribute": "Brand", "present": 10, "total": 10, "coverage_percent": 100}],
        "badges": [{"brand": "Intel", "badge_family": "Core", "coverage_percent": 50, "status": "ok"}],
        "products": [{"retailer": "newegg", "product": "Laptop", "brand": "Intel"}],
    }
    psv = write_psv(tmp_path / "out.psv", tables)
    text = psv.read_text(encoding="utf-8")
    for heading, _key, _sheet, _chart in REPORT_SECTIONS:
        assert heading in text
    assert "Intel|94|94|100|94|93|94|89|99" in text
    assert "AMD|88|88|94|88|65|88|82|100" in text
    assert "NOT_AVAILABLE|Not available for this historical run" in text
    assert "PARTIAL" in text
    excel = write_excel(tmp_path / "out.xlsx", tables)
    from openpyxl import load_workbook

    wb = load_workbook(excel)
    for _heading, _key, sheet, _chart in REPORT_SECTIONS:
        assert sheet in wb.sheetnames
    intel = [
        row
        for row in wb["Brand Compliance"].iter_rows(values_only=True)
        if row and row[0] == "Intel"
    ][0]
    psv_intel = [line for line in text.splitlines() if line.startswith("Intel|94|")][0]
    assert str(intel[-1]) in psv_intel
    wb.close()


def test_report_discovery_groups_latest(tmp_path: Path) -> None:
    day = tmp_path / "2026-08-14"
    day.mkdir()
    (day / "BridgeAI_Report_Run_8_2026-08-14.xlsx").write_bytes(b"xlsx")
    (day / "BridgeAI_Report_Run_8_2026-08-14.psv").write_text("brand|overall\n", encoding="utf-8")
    (day / "BridgeAI_Report_Run_18_2026-08-14.xlsx").write_bytes(b"xlsx")
    (day / "BridgeAI_Report_Run_18_2026-08-14.psv").write_text("brand|overall\n", encoding="utf-8")
    found = discover_reports(root=tmp_path)
    assert [item.run_id for item in found] == [18, 8]
    latest = latest_report(root=tmp_path)
    assert latest is not None
    assert latest.run_id == 18
    assert latest.has_excel and latest.has_psv
    assert "Aug 2026" in latest.display_date


def test_paths_use_run_id_and_date() -> None:
    paths = paths_for_run(18, datetime(2026, 8, 14, 16, 7, tzinfo=timezone.utc))
    assert paths.stem == "BridgeAI_Report_Run_18_2026-08-14"
    assert paths.excel.name.endswith(".xlsx")
    assert paths.psv.name.endswith(".psv")


def test_generate_reports_does_not_mutate_run(session: Session, tmp_path: Path) -> None:
    run = CollectionRun(
        retailer_code="multi",
        country_code="XX",
        run_type="production",
        status=STATUS_SUCCESS,
        started_at=datetime.now(timezone.utc),
        completed_at=datetime.now(timezone.utc),
    )
    session.add(run)
    session.commit()
    generate_reports_for_run(session, run.id, reports_root=tmp_path)
    session.refresh(run)
    assert run.status == STATUS_SUCCESS


def test_collection_status_does_not_invent_success() -> None:
    from dashboard.queries.collection import ComponentStatus

    assert _normalize(None) == "NOT AVAILABLE"
    assert _normalize("SUCCESS") == "SUCCESS"
    assert _normalize("completed") == "SUCCESS"
    assert _normalize("PARTIAL") == "PARTIAL"
    by = {
        "audits": ComponentStatus(component="audits", status="SUCCESS"),
        "newegg": ComponentStatus(component="newegg", status="PARTIAL"),
        "mercadolibre": ComponentStatus(component="mercadolibre", status="SUCCESS"),
    }
    assert _component_status("banners", by) == "NOT AVAILABLE"
    assert _component_status("audits", by) == "SUCCESS"
    assert _component_status("newegg", by) == "PARTIAL"
    assert _component_status("mercadolibre", by) == "SUCCESS"
    assert _component_status("products", by) == "PARTIAL"


def test_report_failure_does_not_fail_collection(session: Session) -> None:
    def _ok(component: str) -> StepResult:
        return StepResult(
            component=component,
            status=STATUS_SUCCESS,
            records_processed=1,
            completed_at=datetime.now(timezone.utc),
        )

    async def _run():
        runner = ProductionRunner(
            session,
            config=OrchestrationConfig(
                product_limit_per_retailer=1,
                search_limit_per_retailer=1,
                stale_running_hours=6,
                concurrent_lock_key=1,
                exit_code_partial=0,
                max_attempts=1,
                base_delay_seconds=0.01,
                component_timeout_seconds=30,
                page_timeout_ms=1000,
            ),
        )
        with (
            patch(
                "collector.orchestration.runner.run_newegg_products",
                new=AsyncMock(return_value=_ok("newegg")),
            ),
            patch(
                "collector.orchestration.runner.run_mercadolibre_products",
                new=AsyncMock(return_value=_ok("mercadolibre")),
            ),
            patch(
                "collector.orchestration.runner.run_audits_step",
                new=AsyncMock(return_value=_ok("audits")),
            ),
            patch(
                "collector.orchestration.runner.run_badges_step",
                new=AsyncMock(return_value=_ok("badges")),
            ),
            patch(
                "collector.orchestration.runner.run_pricing_step",
                new=AsyncMock(return_value=_ok("pricing")),
            ),
            patch(
                "collector.orchestration.runner.run_banners_step",
                new=AsyncMock(return_value=_ok("banners")),
            ),
            patch(
                "collector.orchestration.runner.run_search_step",
                new=AsyncMock(return_value=_ok("search")),
            ),
            patch(
                "collector.orchestration.runner._attempt_reports",
                side_effect=RuntimeError("excel exploded"),
            ),
        ):
            return await runner.run()

    result = asyncio.run(_run())
    assert result.status == STATUS_SUCCESS
    run = session.get(CollectionRun, result.run_id)
    assert run is not None
    assert run.status == STATUS_SUCCESS
    assert result.report_status == "FAILED"


def test_scheduler_scripts_use_venv_python() -> None:
    launcher = Path("scripts/run_scheduled_collection.ps1").read_text(encoding="utf-8")
    setup = Path("scripts/setup_windows_scheduler.ps1").read_text(encoding="utf-8")
    schedule = Path("config/schedule.yaml").read_text(encoding="utf-8")
    docs = Path("docs/windows_scheduler.md").read_text(encoding="utf-8")
    assert ".venv\\Scripts\\python.exe" in launcher
    assert "collector.run" in launcher
    assert "--all" in launcher
    assert "--dry-run" in launcher
    assert "streamlit run" not in launcher.lower()
    assert "WakeToRun" in setup
    assert "IgnoreNew" in setup
    assert "run_scheduled_collection.ps1" in setup
    assert "load_collection_schedule" in setup
    assert "config/schedule.yaml" in setup
    assert "streamlit run" not in setup.lower()
    assert "hours: [8, 14, 20]" in schedule
    assert 'task_name: "BridgeAI - Production Collection"' in schedule
    assert "config/schedule.yaml" in docs
    assert "08:00" in docs and "14:00" in docs and "20:00" in docs
    from collector.orchestration.config import COMPONENTS
    from collector.run import resolve_orchestration_filters, parse_args

    assert COMPONENTS == (
        "newegg",
        "mercadolibre",
        "audits",
        "badges",
        "pricing",
        "banners",
        "search",
    )
    retailers, steps = resolve_orchestration_filters(parse_args(["--all"]))
    assert retailers is None and steps is None


def _production_run(session: Session, *, started_at: datetime) -> CollectionRun:
    run = CollectionRun(
        retailer_code="multi",
        country_code="XX",
        run_type="production",
        status="PARTIAL",
        started_at=started_at,
        completed_at=started_at,
    )
    session.add(run)
    session.flush()
    return run


def _child_pricing_run(session: Session) -> CollectionRun:
    run = CollectionRun(
        retailer_code="newegg",
        country_code="US",
        run_type="pricing",
        status="completed",
        started_at=datetime.now(timezone.utc),
    )
    session.add(run)
    session.flush()
    return run


def _product_with_snapshot(
    session: Session,
    *,
    sku: str,
    title: str,
    run_id: int,
    price: str = "100.00",
) -> Product:
    product = Product(
        retailer_code="newegg",
        country_code="US",
        retailer_sku=sku,
        canonical_url=f"https://example.test/{sku}",
        title=title,
        brand="Intel",
        oem="Asus",
        product_type="notebook",
        category_raw="Gaming Laptops",
        is_active=True,
        last_collection_run_id=run_id,
    )
    session.add(product)
    session.flush()
    session.add(
        ProductSnapshot(
            product_id=product.id,
            collection_run_id=run_id,
            observed_at=datetime.now(timezone.utc),
            title=title,
            brand="Intel",
            oem="Asus",
            product_type="notebook",
            category_raw="Gaming Laptops",
            price_amount=price,
            currency="USD",
        )
    )
    session.add(
        PriceHistory(
            product_id=product.id,
            collection_run_id=run_id,
            observed_at=datetime.now(timezone.utc),
            price_amount=price,
            currency="USD",
        )
    )
    session.add(
        RetailerAudit(
            product_id=product.id,
            collection_run_id=run_id,
            observed_at=datetime.now(timezone.utc),
            retailer_code="newegg",
            country_code="US",
            brand="Intel",
            product_type="notebook",
            check_code="P1",
            result="PASS",
        )
    )
    session.flush()
    return product


def _link_child(session: Session, production: CollectionRun, child: CollectionRun) -> None:
    session.add(
        CollectionRunStep(
            collection_run_id=production.id,
            component="newegg",
            status="SUCCESS",
            details={"child_collection_run_id": child.id, "parent_run_id": production.id},
        )
    )
    session.flush()


def test_parse_run_id_list() -> None:
    assert parse_run_id_list("8,18") == [8, 18]
    assert parse_run_id_list(" 8, 8, 18 ") == [8, 18]


def test_run_id_rejects_nonexistent(session: Session, tmp_path: Path) -> None:
    result = generate_reports_for_run(session, 999, reports_root=tmp_path)
    assert result.status == "FAILED"
    assert "not found" in (result.error_message or "")


def test_observation_run_ids_include_child_from_step_details(session: Session) -> None:
    parent = _production_run(session, started_at=datetime(2026, 8, 14, 13, tzinfo=timezone.utc))
    child = _child_pricing_run(session)
    _link_child(session, parent, child)
    session.commit()
    ids = observation_run_ids_for_run(session, parent.id)
    assert parent.id in ids
    assert child.id in ids


def test_historical_report_does_not_mix_later_run(
    session: Session, tmp_path: Path
) -> None:
    older = _production_run(session, started_at=datetime(2026, 8, 14, 13, tzinfo=timezone.utc))
    older_child = _child_pricing_run(session)
    _link_child(session, older, older_child)
    _product_with_snapshot(
        session,
        sku="OLD-ONLY",
        title="ASUS ROG Gaming Laptop Intel Run One",
        run_id=older_child.id,
        price="111.00",
    )

    newer = _production_run(session, started_at=datetime(2026, 8, 14, 16, tzinfo=timezone.utc))
    newer.status = STATUS_SUCCESS
    newer_child = _child_pricing_run(session)
    _link_child(session, newer, newer_child)
    _product_with_snapshot(
        session,
        sku="NEW-ONLY",
        title="ASUS ROG Gaming Laptop Intel Run Two",
        run_id=newer_child.id,
        price="222.00",
    )
    session.commit()

    commits = {"n": 0}
    original = session.commit

    def _count_commit() -> None:
        commits["n"] += 1
        original()

    session.commit = _count_commit  # type: ignore[method-assign]

    result = generate_reports_for_run(session, older.id, reports_root=tmp_path)
    assert result.ok
    assert commits["n"] == 0
    assert result.excel_path is not None
    assert f"Run_{older.id}_" in result.excel_path.name
    assert result.psv_path is not None
    assert f"Run_{older.id}_" in result.psv_path.name

    from openpyxl import load_workbook

    wb = load_workbook(result.excel_path)
    flat = " ".join(
        str(cell)
        for sheet in wb.worksheets
        for row in sheet.iter_rows(values_only=True)
        for cell in row
    )
    assert "OLD-ONLY" in flat
    assert "NEW-ONLY" not in flat
    assert f"Run_{older.id}_" in result.excel_path.name

    later = generate_reports_for_run(session, newer.id, reports_root=tmp_path)
    assert later.ok
    assert result.excel_path.is_file()
    psv_text = result.psv_path.read_text(encoding="utf-8")
    for heading, _key, _sheet, _chart in REPORT_SECTIONS:
        assert heading in psv_text
    assert "OLD-ONLY" in psv_text
    assert "NEW-ONLY" not in psv_text
    assert "NOT_AVAILABLE" in psv_text
    found = discover_reports(root=tmp_path)
    assert [item.run_id for item in found] == [newer.id, older.id]


def test_historical_skips_component_runs(session: Session, tmp_path: Path) -> None:
    production = _production_run(session, started_at=datetime(2026, 8, 14, 13, tzinfo=timezone.utc))
    child = _child_pricing_run(session)
    session.commit()
    assert list_production_run_ids(session) == [production.id]
    results = generate_historical_production_reports(session, reports_root=tmp_path)
    assert [item.run_id for item in results] == [production.id]
    names = [p.name for p in tmp_path.rglob("BridgeAI_Report_Run_*")]
    assert any(f"Run_{production.id}_" in name for name in names)
    assert not any(f"Run_{child.id}_" in name for name in names)


def test_discovery_missing_psv_and_sort(tmp_path: Path) -> None:
    day = tmp_path / "2026-08-14"
    day.mkdir()
    (day / "BridgeAI_Report_Run_8_2026-08-14.xlsx").write_bytes(b"xlsx")
    (day / "BridgeAI_Report_Run_18_2026-08-14.psv").write_text("brand|overall\n", encoding="utf-8")
    found = discover_reports(root=tmp_path)
    assert [item.run_id for item in found] == [18, 8]
    by_id = {item.run_id: item for item in found}
    assert by_id[8].has_excel and not by_id[8].has_psv
    assert by_id[18].has_psv and not by_id[18].has_excel


def test_attach_run_metadata_keeps_partial(session: Session, tmp_path: Path) -> None:
    run = _production_run(session, started_at=datetime(2026, 8, 14, 15, 35, tzinfo=timezone.utc))
    run.status = "PARTIAL"
    session.commit()
    day = tmp_path / "2026-08-14"
    day.mkdir()
    (day / f"BridgeAI_Report_Run_{run.id}_2026-08-14.xlsx").write_bytes(b"xlsx")
    found = attach_run_metadata(session, discover_reports(root=tmp_path))
    assert found[0].status_label() == "PARTIAL"
    assert found[0].run_id == run.id


def test_one_card_per_run_id_even_with_two_dates(tmp_path: Path) -> None:
    older = tmp_path / "2026-08-13"
    newer = tmp_path / "2026-08-14"
    older.mkdir()
    newer.mkdir()
    (older / "BridgeAI_Report_Run_21_2026-08-13.xlsx").write_bytes(b"old")
    (newer / "BridgeAI_Report_Run_21_2026-08-14.xlsx").write_bytes(b"new")
    (newer / "BridgeAI_Report_Run_21_2026-08-14.psv").write_text("brand|overall\n", encoding="utf-8")
    found = discover_reports(root=tmp_path)
    assert [item.run_id for item in found] == [21]
    assert found[0].has_excel and found[0].has_psv
    latest, previous = split_latest_and_previous(found)
    assert latest is not None and latest.run_id == 21
    assert previous == []


def test_missing_format_does_not_duplicate_run(tmp_path: Path) -> None:
    day = tmp_path / "2026-08-14"
    day.mkdir()
    (day / "BridgeAI_Report_Run_8_2026-08-14.xlsx").write_bytes(b"xlsx")
    (day / "BridgeAI_Report_Run_18_2026-08-14.psv").write_text("psv\n", encoding="utf-8")
    found = discover_reports(root=tmp_path)
    assert [item.run_id for item in found] == [18, 8]
    latest, previous = split_latest_and_previous(found)
    assert latest is not None and latest.run_id == 18
    assert [item.run_id for item in previous] == [8]
    assert latest.has_psv and not latest.has_excel
    assert previous[0].has_excel and not previous[0].has_psv


def test_empty_reports_library(tmp_path: Path) -> None:
    assert discover_reports(root=tmp_path) == []
    latest, previous = split_latest_and_previous([])
    assert latest is None and previous == []


def test_metrics_from_steps_omit_skipped(session: Session) -> None:
    run = _production_run(session, started_at=datetime(2026, 8, 14, 18, 41, tzinfo=timezone.utc))
    session.add_all(
        [
            CollectionRunStep(
                collection_run_id=run.id,
                component="newegg",
                status="SUCCESS",
                records_processed=100,
            ),
            CollectionRunStep(
                collection_run_id=run.id,
                component="mercadolibre",
                status="PARTIAL",
                records_processed=100,
            ),
            CollectionRunStep(
                collection_run_id=run.id,
                component="audits",
                status="SUCCESS",
                records_processed=847,
            ),
            CollectionRunStep(
                collection_run_id=run.id,
                component="banners",
                status="SKIPPED",
                records_processed=0,
            ),
        ]
    )
    session.commit()
    steps = list(session.scalars(select(CollectionRunStep)).all())
    metrics = dict(metrics_from_steps(steps))
    assert metrics["Products"] == 200
    assert metrics["Audits"] == 847
    assert "Banners" not in metrics


def test_reports_view_pairs_excel_psv_on_one_card() -> None:
    from dashboard.queries.reports import DiscoveredReport
    from dashboard.views.reports import _latest_html, _previous_html

    src = Path("dashboard/views/reports.py").read_text(encoding="utf-8")
    assert "generate_reports" not in src
    assert "Download Excel" in src and "Download PSV" in src
    assert "No reports available yet" in src
    report = DiscoveredReport(
        run_id=21,
        date="2026-08-14",
        excel_path=None,
        psv_path=None,
        timestamp=datetime(2026, 8, 14, 18, 41, tzinfo=timezone.utc),
        display_date="14 Aug 2026",
        run_type="production",
        status="PARTIAL",
        retailers=("Newegg", "Mercado Libre"),
        metrics=(("Products", 200), ("Audits", 847)),
    )
    latest = _latest_html(report)
    previous = _previous_html(report)
    assert latest.count("Run 21") == 1
    assert previous.count("Run 21") == 1
    assert "PARTIAL" in latest and "PARTIAL" in previous
    assert "200" in latest and "Products" in latest
    assert "Audits" not in previous

